import os
import random
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font


def run_file(filename):
    print(filename)
    data = pd.read_excel(filename, index_col=0, header=0)
    # data = data.iloc[28:]
    data = data.T
    biolo = data.columns[1:9].tolist()
    # print(biolo)

    bureau_counter = pd.Series([0] * len(biolo), index=biolo)
    delta_counter = pd.Series([0] * len(biolo), index=biolo)

    def get_new_delta_biolo(i):
        n = 2
        l = []
        for b in delta_counter.sort_values().index:
            free_days = data[b].iloc[i:i + 5].isna().sum()
            # print(free_days, b)
            if free_days < 4:
                continue
            l += [b]
            delta_counter[b] += 1
            if len(l) == n:
                break
        return l

    delta_biolo = []

    for i, (day, row) in enumerate(data.iterrows()):
        weekday = row['Weekday']
        if weekday in ['Lun', 'Mer', 'Ven']:
            tasks = ["labo", "labo", "labo", "labo", "PMA (D)", "Sperme (D)"]
        elif weekday in ['Ma', 'Je']:
            tasks = ["labo", "labo", "labo", "labo", "Sperme/PMA (D)",
                     "Sperme (HBW)"]
        else:
            continue
        # print(day)
        alloc = row[biolo]
        taken_tasks = alloc[~pd.isna(alloc)]
        for t in taken_tasks:
            if t in tasks:
                tasks.remove(t)

        free_biolo = alloc[pd.isna(alloc)].index.tolist()

        if not free_biolo:
            continue

        # Delta stuff
        if weekday == 'Lun' or not delta_biolo:
            delta_biolo = get_new_delta_biolo(i)
        # print(delta_biolo)

        delta_tasks = [t for t in tasks if t.endswith('(D)')]
        if weekday == 'Lun':
            delta_tasks = reversed(delta_tasks)

        _delta_biolo = list(
            reversed(delta_biolo)) if weekday == 'Ma' else delta_biolo
        for i, t in enumerate(delta_tasks):
            idx = 0
            while True:
                if idx < len(_delta_biolo):
                    b = _delta_biolo[idx]
                else:
                    b = random.choice(free_biolo)
                if b in free_biolo:
                    free_biolo.remove(b)
                    data.loc[day, b] = t
                    tasks.remove(t)
                    break
                idx += 1

        deficit = len(free_biolo) - len(tasks)
        if deficit > 0:
            for _ in range(deficit):
                for b in bureau_counter.sort_values().index:
                    if b in free_biolo:
                        free_biolo.remove(b)
                        bureau_counter[b] += 1
                        data.loc[day, b] = 'Bureau'
                        break
        elif deficit < 0:
            if pd.isna(data.loc[day, 'Anne VLG']):
                data.loc[day, 'Anne VLG'] = 'labo'
                tasks.pop(tasks.index('labo'))
            if deficit < -1:
                print(f'WARNING: not enough biologists on {day}')

        # shuffle tasks and set free biolo in row:
        n_free = len(free_biolo)
        to_assign = tasks[:n_free]
        # Deterministic daily shuffle based on the date index string for reproducibility
        rs = int(pd.util.hash_pandas_object(pd.Index([str(day)])).iloc[0]) % (
            2 ** 32)
        random.Random(rs).shuffle(to_assign)

        # print(free_biolo)
        # print(tasks)
        # print(to_assign)

        data.loc[day, free_biolo] = to_assign

        # print()

    mask = ~data['Weekday'].isin(['Sam', 'Dim'])
    data.loc[mask, 'Anne VLG'] = data.loc[mask, 'Anne VLG'].fillna('Bureau')
    # print(bureau_counter)

    data = data.T

    # Write and format "labo" cells in blue
    with pd.ExcelWriter(filename, engine='openpyxl', mode="a",
                        if_sheet_exists="replace") as writer:
        data.to_excel(writer, sheet_name='Filled')
        ws = writer.book['Filled']

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1,
                                max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, str):
                    c = cell.value.strip()
                    if c == 'labo':
                        cell.font = Font(color="0000FF")  # Blue
                    elif c in ['Sperme/PMA (D)', 'PMA (D)', 'Sperme (D)']:
                        cell.font = Font(color="FF0000")  # Red
                    elif c in ['Sperme (HBW)']:
                        cell.font = Font(color="00FF00")  # Green

    print(f"Planning mis à jour dans {filename}")


planning_dir = Path('../planning')
filenames = [f for f in os.listdir(planning_dir) if f.endswith('.xlsx')]

for f_name in filenames:
    run_file(planning_dir / f_name)
